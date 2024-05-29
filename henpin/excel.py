import openpyxl as px
from henpin import excelfile

class excel:
    def __init__(self,numList,boxList,datenow,datenext,name):
        self.numList = numList
        self.boxList = boxList
        self.datenow = datenow
        self.datenext = datenext
        self.name = name
    
    def enter_excel(self):
        Excelfile = excelfile.query.get(10)
        HenpinBook ="C:\\Users\\mono2\\OneDrive\\ドキュメント\\デスクトップ\\" + Excelfile.file_name
        wb = px.load_workbook(HenpinBook)

        wsS = wb["書籍"]
        wsZ = wb["雑誌"]
        wsC = wb["雑誌コミック"]
        wsB = wb["文庫"]
        wsO = wb["その他"]

        day = self.datenow.split('/')
        daynext = self.datenext.split('/')
        
        if self.boxList[0] != 0 and self.numList[0] != "":
            numList0 = int(self.numList[0])
            for a in range(self.boxList[0]):
                wsS.cell(row=numList0 + 3 + a,column=6).value = day[0] + "月" + day[1] + "日"
                wsS.cell(row=numList0 + 3 + a,column=7).value = daynext[0] + "月" + daynext[1] + "日"
                wsS.cell(row=numList0 + 3 + a,column=8).value = self.name

        if self.boxList[1] != 0 and self.numList[1] != "":
            numList1 = int(self.numList[1])
            for b in range(self.boxList[1]):
                wsZ.cell(row=numList1 + 3 + b,column=6).value = day[0] + "月" + day[1] + "日"
                wsZ.cell(row=numList1 + 3 + b,column=7).value = daynext[0] + "月" + daynext[1] + "日"
                wsZ.cell(row=numList1 + 3 + b,column=8).value = self.name
        
        if self.boxList[2] != 0 and self.numList[2] != "":
            numList2 = int(self.numList[2])
            for c in range(self.boxList[2]):
                wsC.cell(row=numList2 + 3 + c,column=6).value = day[0] + "月" + day[1] + "日"
                wsC.cell(row=numList2 + 3 + c,column=7).value = daynext[0] + "月" + daynext[1] + "日"
                wsC.cell(row=numList2 + 3 + c,column=8).value = self.name
        
        if self.boxList[3] != 0 and self.numList[3] != "":
            numList3 = int(self.numList[3])
            for d in range(self.boxList[3]):
                wsB.cell(row=numList3 + 3 + d,column=6).value = day[0] + "月" + day[1] + "日"
                wsB.cell(row=numList3 + 3 + d,column=7).value = daynext[0] + "月" + daynext[1] + "日"
                wsB.cell(row=numList3 + 3 + d,column=8).value = self.name
        
        if self.boxList[4] != 0 and self.numList[4] != "":
            numList4 = int(self.numList[4])
            for e in range(self.boxList[4]):
                wsO.cell(row=numList4 + 3 + e,column=6).value = day[0] + "月" + day[1] + "日"
                wsO.cell(row=numList4 + 3 + e,column=7).value = daynext[0] + "月" + daynext[1] + "日"
                wsO.cell(row=numList4 + 3 + e,column=8).value = self.name

        wb.save(HenpinBook)